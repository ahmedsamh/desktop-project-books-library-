# image package
import tkinter as tk
from PIL import ImageTk, Image
import requests
from io import BytesIO

from tkinter import *

from datetime import date

from tkinter import filedialog
from tkinter import messagebox

from PIL import Image, ImageTk

import os

from tkinter.ttk import Combobox

import openpyxl
import xlrd

from openpyxl import Workbook

import pathlib
import pymongo
book = pymongo.MongoClient("mongodb://localhost:27017/")
db = book["book_app"]
collection = db["book"]
studentCollection = db["students"]
background = "#132E3C"
framebg = "#EDEDED"

framefg = "#062830"
root = Tk()
root.title("Book System")

root.geometry("1250x700+210+100")

root.config(bg=background)


# Function to load and display the online image

print('****************')
print(studentCollection.find_one({"_id": 2}))
# Exit


def Exit():
    root.destroy()
# Shoe image#########333


def showimage():
    global filename
    global img_book
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetype=(("JPG File", "*.jpg"),
                                                                               ("PNG File",
                                                                                "*.png"),
                                                                               ("ALL files", "*.txt")))
    img_book = (Image.open(filename))
    resized_image = img_book.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lb_book.config(image=photo2)
    lb_book.image = photo2

# registration NO #######################3


def registration_no():
    file = openpyxl.load_workbook('Book_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

# Clear ########3


def Clear():
    global img
    global img_book
    # Name.set('')
    Student_Id.set('')
    Book_Name.set('')
    author.set('')
    Description.set('')
    Category.set('')
    publisher.set('')
    page_count.set('')
    average.set('')
    rating.set('')
    subtitle.set('')
    saveButton.config(state='normal')

    Book_Photo = Frame(obj2, bd=3, bg="black", width=230,
                       height=180, relief=GROOVE)
    Book_Photo.place(x=30, y=10)

    img_book = PhotoImage(
        file="C:/Users/amrka/OneDrive/Desktop/python/python/images/profile2.png")

 # book
    lb_book = Label(Book_Photo, bg="black",
                    image=img_book, width=230, height=180)
    lb_book.place(x=0, y=0)


################ SAVE ##############


def Exist(id):
    count = collection.count_documents({'_id': id})
    return count > 0


def finduser(id):
    count = studentCollection.count_documents({'_id': id})
    return count > 0


def Save():
    id_book = Registration.get()
    _Book_name = Book_Name.get()
    _author = author.get()
    _Description = Description.get()
    _cate = Category.get()
    _publisher = publisher.get()
    _page_count = page_count.get()
    _subtitle = subtitle_entry.get()
    _average_rating = average_entry.get()
    _rating = rating_entry.get()
    _userid = userId.get()

    if _Book_name == "" or _subtitle == "" or _author == "" or _rating == "" or _average_rating == "" or _Description == "" or _cate == "" or _publisher == "" or _page_count == "":
        messagebox.showerror("error", "Few Data is Missing")
    elif (Exist(id_book)):
        messagebox.showerror("error", "book exists")
    elif (userId != ''):
        if (finduser(_userid)):
            book = {
                '_id': id_book,
                "title": _Book_name,
                "subtitle": _subtitle,
                "authors": _author,
                "description": _Description,
                "categories": _cate,
                "published_year": _publisher,
                "average_rating": _average_rating,
                "num_pages": _page_count,
                "ratings_count": _rating,
                "UserId": _userid

            }
            collection.insert_one(book)

            olduser = studentCollection.find_one({"_id": _userid})
            oldbook = list(olduser['Books'])
            oldbook.append(_Book_name)

            studentCollection.update_one(
                {"_id": _userid}, {"$set": {"Books": oldbook}})

            messagebox.showinfo("info", "Sucessfully data entered!!!")
            Clear()
        else:
            messagebox.showerror("error", "User Not Found!!!")

    else:
        book = {
            '_id': id_book,
            "title": _Book_name,
            "subtitle": _subtitle,
            "authors": _author,
            "description": _Description,
            "categories": _cate,
            "published_year": _publisher,
            "average_rating": _average_rating,
            "num_pages": _page_count,
            "ratings_count": _rating

        }
        collection.insert_one(book)

        messagebox.showinfo("info", "Sucessfully data entered!!!")
        Clear()  # clear entry box and image section
    # it will recheck registration no. and reissue new no.
       # Registration.set(id_book + 1)

############# SEArch ######################


def search():
    text = Search.get()
    search_result = collection.find_one({'_id': text})
    if search_result is not None:
        print(search_result['title'])
        # Access the fields of the document
        Registration.set(search_result['_id'])
        Book_Name.set(search_result['title'])
        author.set(search_result['authors'])
        Description.set(search_result['description'])
        publisher.set(search_result['published_year'])
        Category.set(search_result['categories'])
        page_count.set(search_result['num_pages'])
        average.set(search_result['average_rating'])
        rating.set(search_result['ratings_count'])
        subtitle.set(search_result['subtitle'])
        load_image_url(search_result['thumbnail'])

        # Book_Photo = Frame(obj2, bd=3, bg="black", width=230,
        # height=180, relief=GROOVE)
        # Book_Photo.place(x=30, y=10)
        # img_book = PhotoImage(
        # file="C:/Users/amrka/OneDrive/Desktop/python/python/images/profile2.png")  # book
        # lb_book = Label(Book_Photo, bg="black", image=load_image_url(search_result['thumbnail']), width=230, height=180)
        # lb_book.place(x=0, y=0)

        saveButton.configure(state='disable')
        updateButton = Button(obj2, text="update", width=10, height=2,
                              font="arial 12 bold", bg="#829DAB", command=Update)
        updateButton.place(x=550, y=200)
    else:
        messagebox.showerror("error", "book is`nt Here")
        # Handle the case where nodocument is found
        Clear()

    # Backend fetch data


# Update ############3
def Update():
    id_book = Search.get()
    _Book_name = Book_Name.get()
    _author = author.get()
    _Description = Description.get()
    _cate = Category.get()
    _publisher = publisher.get()
    _page_count = page_count.get()
    _subtitle = subtitle_entry.get()
    _average_rating = average_entry.get()
    _rating = rating_entry.get()

    if _Book_name == "" or _subtitle == "" or _author == "" or _rating == "" or _average_rating == "" or _Description == "" or _cate == "" or _publisher == "" or _page_count == "":
        messagebox.showerror("error", "Few Data is Missing")

    else:
        book = {
            '_id': id_book,
            "title": _Book_name,
            "subtitle": _subtitle,
            "authors": _author,
            "description": _Description,
            "categories": _cate,
            "published_year": _publisher,
            "average_rating": _average_rating,
            "num_pages": _page_count,
            "ratings_count": _rating,
        }
        collection.update_one({"_id": id_book}, {"$set": book})

        messagebox.showinfo("info", "Updated Sucessfully data !!!")
        Clear()
    # clear entry box and image section
    # it will recheck registration no. and reissue new no.
    # Registration.set(id_book + 1)


####
def delete():
    id_delele = Search.get()
    if (Exist(id_delele)):

        collection.delete_one({'_id': id_delele})
        Clear()
        saveButton.configure(state='disable')
        messagebox.showinfo("info", "Sucessfully data Deleted!!!")

    else:
        messagebox.showerror("error", "Book Not Found")

    # Clear()
    # saveButton.configure(state='disable')
    # print(id_delele)


# stop frames
Label(root, text="BIG DATA PROJECT ", width=10, height=3,
      bg="#707F82",  font="arial 10 bold").pack(side=TOP, fill=X)
Label(root, text="INSERT BOOK DATA", width=10, height=2, bg="#94A3A2",
      fg='#fff', font="arial 20 bold").pack(side=TOP, fill=X)
Label(root, text="BY : Ahmed Sameh , Ahmed Khaled , Amr Kasban , Ziad Ahmed , Salwa Ashraf ",
      bg="#707F82",  font="arial 15 bold").pack(side=BOTTOM, fill=X)

Search = StringVar()
Entry(root, textvariable=Search, width=15,
      bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(
    file="C:/Users/amrka/OneDrive/Desktop/python/python/images/search.png")
Srch = Button(root, text="Search", image=imageicon3, width=30,
              height=30, bg='#94A3A2', font="arial 13 bold", command=search)
Srch.place(x=1060, y=70)

imageIconDelete = PhotoImage(
    file="C:/Users/amrka/OneDrive/Desktop/python/python/images/profile2.png")  # dele
delete = Button(root, text="Search", image=imageIconDelete, width=30,
                height=30, bg='#94A3A2', font="arial 13 bold", command=delete)
delete.place(x=780, y=70)


# Registration and Date


Label(root, text="Date:", font="arial 13",
      fg=framebg, bg=background).place(x=500, y=150)

Date = StringVar()
today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)
Date.set(d1)

obj = LabelFrame(root, text="Student's Details", font=20, bd=2,
                 width=1200, bg=framebg, fg=framefg, height=120, relief=GROOVE)
obj.place(x=30, y=200)


Label(obj, text="First Name:", font="arial 13",
      bg=framebg, fg=framefg). place(x=30, y=20)
Label(obj, text="Student_Id: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=300, y=50)

Label(obj, text="Last Name:", font="arial 13",
      bg=framebg, fg=framefg). place(x=30, y=50)
Label(obj, text="Student_Id: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=300, y=50)


firstname = StringVar()
firstname_entry = Entry(obj, textvariable=firstname, width=20, font="arial 10")
firstname_entry.place(x=120, y=20)

Lastname = StringVar()
Lastname_entry = Entry(obj, textvariable=Lastname, width=20, font="arial 10")
Lastname_entry.place(x=120, y=50)

Student_Id = IntVar()
Student_Id_entry = Entry(obj, textvariable=Student_Id,
                         width=15, font="arial 10")
Student_Id_entry.place(x=410, y=50)


def findallbooktitle():
    res = list(collection.find({}, {"title": 1, "_id": 0}))
    titles = []
    for x in res:
        titles.append(x['title'])
    return titles


Class = Combobox(obj, values=findallbooktitle(),
                 font='Roboto 10', width=17, state="r")
Class.place(x=410, y=15)
Class.set("SELECT BOOK")

from bson.code import Code



# Call the function to perform the mapReduce operation on the title field
def insert_user():
    user_id = Student_Id.get()
    first_name = firstname.get()
    last_name = Lastname.get()
    choosebook = Class.get()
    if user_id == "" or first_name == "" or last_name == "" or choosebook == "SELECT BOOK":
        messagebox.showerror("error", "Fill Data")
    else:
        user = {
            '_id': user_id,
            'firstname': first_name,
            'lastname': last_name,
            'Books': [choosebook]
        }
        studentCollection.insert_one(user)
        messagebox.showinfo("info", "User Inserted sucessfuly data !!!")


def clearUser():
    Student_Id.set('')
    firstname.set('')
    Lastname.set('')
    Class.set('')


def updateUser():
    user_id = Student_Id.get()
    first_name = firstname.get()
    last_name = Lastname.get()
    choosebook = Class.get()
    if user_id == "" or first_name == "" or last_name == "":
        messagebox.showerror("error", "Fill Data")
    else:
        user = {
            '_id': user_id,
            'firstname': first_name,
            'lastname': last_name,

        }
        studentCollection.update_one({"_id": user_id}, {"$set": user})
        messagebox.showinfo("info", "Updated SuccessFully data !!!")
        clearUser()


# search for specific user
def searchForUser():
    global delete_button
    id = UserIdSearch.get()
    if (finduser(id)):
        result = studentCollection.find_one({"_id": id})
        Student_Id.set(result['_id'])
        firstname.set(result['firstname'])
        Lastname.set(result['lastname'])
        insert_button = Button(obj, text="Updata", width=15, height=1, font="arial 12 bold",
                               bg="#829DAB", command=updateUser)
        insert_button.place(x=700, y=50)

        # show delete

        delete_button = Button(obj, text="Delete",  width=15, height=1, font="arial 12 bold",
                               bg="#829DAB", command=deleteUser)
        delete_button.place(x=700, y=10)
        insert_button.pack()
        insert_button.pack_forget()

        listbooks = StringVar()
        listbooks_entry = Entry(
            obj, textvariable=listbooks, width=55,  font="arial 10")
        listbooks_entry.place(x=900, y=0)
        listbooks.set(result['Books'])

    else:
        messagebox.showerror('error', 'User Not Found')
        clearUser()
        delete_button.pack()

# Hide the button
        delete_button.pack_forget()
        listbooks_entry.pack()

# Hide the button
        listbooks_entry.pack_forget()
# get all book greate than 200


def fetchBookGte():
    return collection.aggregate(
        [{
         "$match":
         {
             "num_pages": {"$gte": "800"}
         }
         }
         ])



def getDocumentWithMaxAverageRating():
    pipeline = [
        {
            "$group": {
                "_id": None,
                "max_average_rating": {"$max": "$average_rating"},
            }
        },
        
    ]

    result = collection.aggregate(pipeline)

    for doc in result:
        max_average_rating = doc["max_average_rating"]
        max_rating_doc = collection.find_one(
            {"average_rating": str(max_average_rating)})
        return max_rating_doc



# Usage example
max_rating_doc = getDocumentWithMaxAverageRating()




def DocumentCount():
    result = collection.aggregate([
        {
            "$match": {
                "num_pages": {"$gte": "100"}
            }
        },
        {
            "$group": {
                "_id": None,
                "count": {"$sum": 1}
            }
        }
    ])

    # Extract the count from the result
    count = 0
    for doc in result:
        count = doc["count"]
        break

    return count


def sortAndLimitDataByTitle():
    pipeline = [
        {"$sort": {"title": 1}},  # Sort in ascending order by title
        {"$limit": 5}  # Limit the output to 5 documents
    ]

    result = collection.aggregate(pipeline)
    for doc in result:
        print(doc)

# Call the function to sort and limit the data by title


def createTitleIndex():
    collection.create_index("title")
    print("Index created successfully.")

def sortDataByTitle():
    pipeline = [
        {"$sort": {"title": 1}}  # Sort in ascending order by title
    ]

    result = collection.aggregate(pipeline)
    for doc in result:
        print(doc)

# Call the function to sort the data by title



def deleteUser():
    id_delele = UserIdSearch.get()
    if (Exist(id_delele)):

        studentCollection.delete_one({'_id': id_delele})
        clearUser()

        messagebox.showinfo("info", "Sucessfully data Deleted!!!")

    else:
        messagebox.showerror("error", "Book Not Found")


print(studentCollection.find_one({"_id": 2}))

# insert
Label(obj, text="Insert Data:", font="arial 13",
      bg=framebg, fg=framefg). place(x=600, y=50)
# dro
insert_button = Button(obj, text="Insert", width=15, height=1, font="arial 12 bold",
                       bg="#829DAB", command=insert_user)
insert_button.place(x=700, y=50)


obj2 = LabelFrame(root, text="Book's Details", font=20, bd=2,
                  width=900, bg=framebg, fg=framefg, height=330, relief=GROOVE)
obj2.place(x=30, y=340)


Book_Photo = Frame(obj2, bd=3, bg="black", width=230,
                   height=180, relief=GROOVE)
Book_Photo.place(x=30, y=10)

img_book = PhotoImage(
    file="C:/Users/amrka/OneDrive/Desktop/python/python/images/profile2.png")

# book
lb_book = Label(Book_Photo, bg="black", image=img_book, width=230, height=180)
lb_book.place(x=0, y=0)


Label(obj2, text=" Title:", font="arial 13",
      bg=framebg, fg=framefg). place(x=270, y=20)

Label(obj2, text="UserId:", font="arial 13",
      bg=framebg, fg=framefg). place(x=270, y=50)

Label(obj2, text="Author: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=270, y=110)

Label(obj2, text="Description: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=270, y=150)
Label(obj2, text="Category: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=270, y=200)
Label(obj2, text="publisher: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=550, y=50)
Label(obj2, text="page count: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=550, y=100)

Label(obj2, text="BOOk Id:", font="arial 13",
      fg=framefg, bg=framebg).place(x=250, y=270)
Label(obj2, text="Subtitle: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=270, y=80)

Label(obj2, text="Rating: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=550, y=150)

Label(obj2, text="Average rating: ", font="arial 13",
      bg=framebg, fg=framefg).place(x=550, y=10)


# register label
Registration = IntVar()
Registration.set(1)
reg_entry = Entry(obj2, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=370, y=270)

imageicon = PhotoImage(
    file="C:/Users/amrka/OneDrive/Desktop/python/python/images/search.png")
Srch = Button(root, text="Search", image=imageicon, width=30,
              height=30, bg='#94A3A2', font="arial 13 bold", command=searchForUser)
Srch.place(x=320, y=150)
UserIdSearch = IntVar()
Registration.set(1)
UserIdSearch_entry = Entry(
    root, textvariable=UserIdSearch, width=15, font="arial 10")
UserIdSearch_entry.place(x=370, y=150)


userId = IntVar()
userId_entry = Entry(obj2, textvariable=userId, width=20, font="arial 10")
userId_entry.place(x=370, y=50)


Book_Name = StringVar()
BookN_entry = Entry(obj2, textvariable=Book_Name, width=20, font="arial 10")
BookN_entry.place(x=370, y=20)

subtitle = StringVar()
subtitle_entry = Entry(obj2, textvariable=subtitle, width=20, font="arial 10")
subtitle_entry.place(x=370, y=80)


rating = StringVar()
rating_entry = Entry(obj2, textvariable=rating, width=20, font="arial 10")
rating_entry.place(x=650, y=150)


average = StringVar()
average_entry = Entry(obj2, textvariable=average, width=20, font="arial 10")
average_entry.place(x=680, y=18)


author = StringVar()
Author_entry = Entry(obj2, textvariable=author, width=20, font="arial 10")
Author_entry.place(x=370, y=110)

Description = StringVar()
Description_entry = Entry(
    obj2, textvariable=Description, width=20, font="arial 10")
Description_entry.place(x=370, y=160)

Category = StringVar()
Category_entry = Entry(obj2, textvariable=Category, width=20, font="arial 10")
Category_entry.place(x=370, y=210)

publisher = StringVar()
publisher_entry = Entry(obj2, textvariable=publisher,
                        width=20, font="arial 10")
publisher_entry.place(x=650, y=50)

page_count = StringVar()
page_count_entry = Entry(obj2, textvariable=page_count,
                         width=20, font="arial 10")
page_count_entry.place(x=650, y=100)


# Save button book data
saveButton = Button(obj2, text="Save", width=10, height=2,
                    font="arial 12 bold", bg="#829DAB", command=Save)
saveButton.place(x=690, y=200)
# button


Button(root, text="Reset", width=22, height=2, font="arial 12 bold",
       bg="#829DAB", command=Clear).place(x=1000, y=430)
Button(root, text="Exit", width=22, height=2, font="arial 12 bold",
       bg="#829DAB", command=Exit).place(x=1000, y=510)


def load_image_url(imageLink):
    url = imageLink  # Replace with your image URL

    # Fetch the image from the URL
    response = requests.get(url)
    image_data = response.content

    # Load the image using PIL
    image = Image.open(BytesIO(image_data))
    # Resize the image as per your requirement
    image = image.resize((240, 220))

    # Create a Tkinter-compatible image object
    tk_image = ImageTk.PhotoImage(image)

    # Create a label to display the image

    # label = tk.Label(root, image=tk_image)
    # label.pack()

    Book_Photo = Frame(obj2, bd=3, bg="black", width=230,
                       height=180, relief=GROOVE)
    Book_Photo.place(x=30, y=10)
    # img_book = PhotoImage(
    # file="C:/Users/amrka/OneDrive/Desktop/python/python/images/profile2.png")  # book
    lb_book = Label(Book_Photo, bg="black",
                    image=tk_image, width=230, height=180)
    lb_book.place(x=0, y=0)

    # Keep a reference to the image to prevent it from being garbage collected
    lb_book.image_names = tk_image


root.mainloop()
